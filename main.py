"""
Email Digest Bot: Gmail + Auto.ru API → compact Telegram digest.
Sources: Excel from email (no-reply@business.auto.ru) + POST /comeback API.
Offers are merged by offer_id, deduplicated, and sent as a single digest.
"""
import imaplib
import email
import os
import json
import urllib.request
import urllib.error
import time
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from email.header import decode_header
from collections import defaultdict
from typing import Optional

from dotenv import load_dotenv
import openpyxl

load_dotenv()

GMAIL_USER = os.getenv("GMAIL_USER")
GMAIL_APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD")
IMAP_SERVER = "imap.gmail.com"
SENDER_FILTER = "no-reply@business.auto.ru"

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID")

VERTIS_SESSION_ID = os.getenv("VERTIS_SESSION_ID")
COMEBACK_API_URL = "https://apiauto.ru/1.0/comeback"

DOWNLOAD_DIR = os.path.join(os.path.dirname(__file__), "downloads")
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

STATE_FILE = os.path.join(os.path.dirname(__file__), "processed_ids.json")


# ── Data model ────────────────────────────────────────────────────

@dataclass
class ComebackOffer:
    offer_id: str           # "1131420679-5346d8a6"
    brand: str              # "TOYOTA"
    model: str              # "CAMRY"
    salon: str              # "EKT"
    category: str           # "not_purchased" | "back_on_sale"
    mobile_url: str
    source: str             # "email" | "api" | "both"
    price: Optional[int] = None
    mileage: Optional[int] = None
    year: Optional[int] = None


# ── State management (auto-migrating) ────────────────────────────

def load_state():
    """Load state dict. Auto-migrate from old list format."""
    try:
        with open(STATE_FILE, "r") as f:
            data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        data = None

    # Old format: plain list of email message IDs
    if isinstance(data, list):
        print("[State] Migrating from list → dict format")
        return {
            "email_message_ids": data,
            "api_offer_ids": [],
            "api_last_fetch": None,
        }

    if isinstance(data, dict):
        # Ensure all keys exist
        data.setdefault("email_message_ids", [])
        data.setdefault("api_offer_ids", [])
        data.setdefault("api_last_fetch", None)
        return data

    return {"email_message_ids": [], "api_offer_ids": [], "api_last_fetch": None}


def save_state(state):
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, ensure_ascii=False)


# ── Helpers (unchanged) ──────────────────────────────────────────

def decode_header_value(value):
    if not value:
        return ""
    parts = decode_header(value)
    result = []
    for part, encoding in parts:
        if isinstance(part, bytes):
            result.append(part.decode(encoding or "utf-8", errors="replace"))
        else:
            result.append(part)
    return " ".join(result)


def extract_url(val):
    """Extract URL from =HYPERLINK(...) formula."""
    if isinstance(val, str) and val.startswith("=HYPERLINK"):
        parts = val.split('"')
        if len(parts) >= 2:
            return parts[1]
    return val if isinstance(val, str) else ""


def extract_offer_id(url):
    """Extract offer ID from auto.ru URL like '1131420679-5346d8a6'."""
    if not url:
        return ""
    parts = url.rstrip("/").split("/")
    if parts:
        last = parts[-1]
        if "-" in last and any(c.isdigit() for c in last):
            return last
    return ""


def normalize_offer_id(oid):
    """Normalize offer_id for matching: lowercase, strip, no trailing slash."""
    return oid.lower().strip().rstrip("/")


def make_mobile_link(url):
    """Convert auto.ru URL to m.auto.ru."""
    if not url:
        return ""
    return url.replace("https://auto.ru/", "https://m.auto.ru/")


SALON_SHORT = {
    "июль екб совхозная": "EKT",
    "июль екб металлургов": "EKT",
    "июль екб базовый": "EKT",
    "июль екб": "EKT",
    "июль екб  рынок год": "EKT",
    "июль екб выезд": "EKT",
    "июль члб копейское": "CHL",
    "июль члб": "CHL",
    "июль крд": "KRD",
    "omoda новые": "OMODA",
}


def short_salon(name):
    """Convert salon name to short code."""
    if not name:
        return "???"
    lower = name.lower().strip()
    for key, code in SALON_SHORT.items():
        if key in lower:
            return code
    return name.split()[0][:3].upper() if name else "???"


def col_index(headers, *names):
    """Find column index by name (case-insensitive)."""
    h_lower = {str(h).lower(): i for i, h in enumerate(headers) if h}
    for name in names:
        if name in h_lower:
            return h_lower[name]
    return None


def get(row, idx):
    """Safe get from row by index."""
    if idx is not None and idx < len(row):
        return row[idx]
    return None


# ── Gmail ────────────────────────────────────────────────────────

def fetch_today_emails(state):
    """Fetch today's emails, return (results, new_message_ids)."""
    processed = set(state.get("email_message_ids", []))
    print(f"[Gmail] Connecting as {GMAIL_USER}... ({len(processed)} already processed)")
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(GMAIL_USER, GMAIL_APP_PASSWORD)
    mail.select("inbox")

    today = datetime.now().strftime("%d-%b-%Y")
    status, messages = mail.search(
        None, f'(FROM "{SENDER_FILTER}" SINCE {today})'
    )

    if status != "OK" or not messages[0]:
        print("[Gmail] No emails found today")
        mail.logout()
        return [], set()

    email_ids = messages[0].split()
    print(f"[Gmail] Found {len(email_ids)} emails today")

    results = []
    new_ids = set()
    for eid in email_ids:
        status, data = mail.fetch(eid, "(RFC822)")
        if status != "OK":
            continue

        msg = email.message_from_bytes(data[0][1])
        msg_id = msg.get("Message-ID", eid.decode())
        if msg_id in processed:
            print(f"[Gmail] Skip (already sent): {msg_id}")
            continue

        subject = decode_header_value(msg["Subject"])
        files = []

        for part in msg.walk():
            filename = part.get_filename()
            if not filename:
                continue
            decoded_name = decode_header_value(filename)
            if decoded_name.endswith((".xlsx", ".xls", ".csv")):
                safe_name = decoded_name.replace("/", "_").replace("\\", "_")
                filepath = os.path.join(DOWNLOAD_DIR, safe_name)
                with open(filepath, "wb") as f:
                    f.write(part.get_payload(decode=True))
                files.append(filepath)

        if files:
            results.append((subject, files))
            new_ids.add(msg_id)

    mail.logout()
    return results, new_ids


# ── Excel → ComebackOffer[] ──────────────────────────────────────

def parse_not_purchased(filepath):
    """Parse 'Не выкупленные' Excel → list[ComebackOffer]."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    offers = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = rows[0]
        data = rows[1:]
        if not data:
            continue

        if "совпадения" in sheet_name.lower():
            i_brand = col_index(headers, "марка")
            i_model = col_index(headers, "модель")
            i_salon = col_index(headers, "автосалон")
            i_link = col_index(headers, "ссылка на объявление")

            seen = set()
            for row in data:
                url = extract_url(get(row, i_link) or "")
                offer_id = extract_offer_id(url)
                if not offer_id or offer_id in seen:
                    continue
                seen.add(offer_id)

                offers.append(ComebackOffer(
                    offer_id=offer_id,
                    brand=str(get(row, i_brand) or "").upper(),
                    model=str(get(row, i_model) or "").upper().replace(" ", "_"),
                    salon=short_salon(get(row, i_salon)),
                    category="not_purchased",
                    mobile_url=make_mobile_link(url),
                    source="email",
                ))

    wb.close()
    return offers


def parse_back_on_sale(filepath):
    """Parse 'Снова в продаже' Excel → list[ComebackOffer]."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    offers = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = rows[0]
        data = rows[1:]
        if not data:
            continue

        if "найденные" in sheet_name.lower():
            i_brand = col_index(headers, "марка")
            i_model = col_index(headers, "модель")
            i_salon = col_index(headers, "автосалон")
            i_link = col_index(headers, "ссылка на объявление")

            seen = set()
            for row in data:
                url = extract_url(get(row, i_link) or "")
                offer_id = extract_offer_id(url)
                if not offer_id or offer_id in seen:
                    continue
                seen.add(offer_id)

                offers.append(ComebackOffer(
                    offer_id=offer_id,
                    brand=str(get(row, i_brand) or "").upper(),
                    model=str(get(row, i_model) or "").upper().replace(" ", "_"),
                    salon=short_salon(get(row, i_salon)),
                    category="back_on_sale",
                    mobile_url=make_mobile_link(url),
                    source="email",
                ))

    wb.close()
    return offers


# ── Auto.ru Comeback API ─────────────────────────────────────────

# Salon ID → short code mapping for API responses
SALON_ID_SHORT = {
    "ekb": "EKT",
    "chel": "CHL",
    "krd": "KRD",
}


def _api_request(body, attempt=1):
    """Single API request. Returns parsed JSON or None."""
    data = json.dumps(body).encode("utf-8")
    req = urllib.request.Request(
        COMEBACK_API_URL,
        data=data,
        headers={
            "Content-Type": "application/json",
            "x-session-id": VERTIS_SESSION_ID,
            "X-Authorization": VERTIS_SESSION_ID,
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        body_text = ""
        try:
            body_text = e.read().decode()
        except Exception:
            pass
        print(f"[API] HTTP {e.code} (attempt {attempt}): {body_text[:300]}")
        if e.code == 401:
            print("[API] 401 Unauthorized — VERTIS_SESSION_ID expired or invalid")
            return None
        if attempt < 2 and e.code in (429, 500, 502, 503):
            time.sleep(3)
            return _api_request(body, attempt + 1)
        return None
    except Exception as e:
        print(f"[API] Error (attempt {attempt}): {e}")
        if attempt < 2:
            time.sleep(3)
            return _api_request(body, attempt + 1)
        return None


def _parse_api_offer(item):
    """Parse single API comeback item → ComebackOffer or None."""
    offer = item.get("offer", {})
    offer_id = offer.get("id", "")
    if not offer_id:
        return None

    # Category: not_purchased or back_on_sale
    comeback_type = item.get("comeback_type", "")
    if comeback_type == "NOT_PURCHASED":
        category = "not_purchased"
    elif comeback_type == "BACK_ON_SALE":
        category = "back_on_sale"
    else:
        category = "not_purchased"  # default

    # Brand / model
    car_info = offer.get("car_info", {})
    brand = car_info.get("mark_info", {}).get("name", "").upper()
    model = car_info.get("model_info", {}).get("name", "").upper().replace(" ", "_")

    # Salon
    salon_info = offer.get("salon", {})
    salon_name = salon_info.get("code", "") or salon_info.get("name", "")
    salon = "???"
    if salon_name:
        sn_lower = salon_name.lower()
        for key, code in SALON_ID_SHORT.items():
            if key in sn_lower:
                salon = code
                break
        if salon == "???":
            salon = short_salon(salon_name)

    # Price
    price_info = offer.get("price_info", {})
    price = price_info.get("price") or price_info.get("RUR")
    if price:
        price = int(price)
    else:
        price = None

    # Mileage, year
    state_info = offer.get("state", {})
    mileage = state_info.get("mileage")
    if mileage:
        mileage = int(mileage)
    else:
        mileage = None

    documents = offer.get("documents", {})
    year = documents.get("year")
    if year:
        year = int(year)
    else:
        year = None

    # URL
    url = offer.get("url", "")
    if not url and offer_id:
        section = offer.get("section", "used").lower()
        b = brand.lower()
        m = model.lower().replace("_", "-")
        url = f"https://auto.ru/cars/{section}/sale/{b}/{m}/{offer_id}/"
    mobile_url = make_mobile_link(url) if url else f"https://m.auto.ru/cars/used/sale/{brand.lower()}/{model.lower().replace('_', '-')}/{offer_id}/"

    return ComebackOffer(
        offer_id=offer_id,
        brand=brand,
        model=model,
        salon=salon,
        category=category,
        mobile_url=mobile_url,
        source="api",
        price=price,
        mileage=mileage,
        year=year,
    )


def fetch_api_comeback():
    """Fetch comeback offers from Auto.ru API for yesterday. Returns list[ComebackOffer]."""
    if not VERTIS_SESSION_ID:
        print("[API] VERTIS_SESSION_ID not set — skipping API")
        return []

    # Yesterday 00:00 → today 00:00 (Moscow, UTC+3)
    msk = timezone(timedelta(hours=3))
    now_msk = datetime.now(msk)
    yesterday_start = now_msk.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=1)
    today_start = yesterday_start + timedelta(days=1)

    date_from = int(yesterday_start.timestamp() * 1000)
    date_to = int(today_start.timestamp() * 1000)

    print(f"[API] Fetching comeback {yesterday_start.strftime('%d.%m.%Y')} → {today_start.strftime('%d.%m.%Y')}")

    all_offers = []
    page = 1
    page_size = 50

    while True:
        body = {
            "filter": {
                "creation_date_from": str(date_from),
                "creation_date_to": str(date_to),
            },
            "page": page,
            "page_size": page_size,
        }

        result = _api_request(body)
        if not result:
            if page == 1:
                print("[API] Failed to fetch — continuing with email-only")
            break

        items = result.get("offers", []) or result.get("items", []) or []
        if not items:
            if page == 1:
                print("[API] No comeback offers found for yesterday")
            break

        for item in items:
            offer = _parse_api_offer(item)
            if offer:
                all_offers.append(offer)

        total = result.get("pagination", {}).get("total_count", 0)
        fetched = page * page_size
        print(f"[API] Page {page}: {len(items)} items (total: {total})")

        if fetched >= total or len(items) < page_size:
            break
        page += 1

    print(f"[API] Total: {len(all_offers)} offers")
    return all_offers


# ── Merge ─────────────────────────────────────────────────────────

def merge_offers(email_offers, api_offers):
    """Merge email + API offers by offer_id. Returns list[ComebackOffer]."""
    # Index email offers by normalized ID
    merged = {}
    for o in email_offers:
        key = normalize_offer_id(o.offer_id)
        merged[key] = o

    # Merge API offers
    for o in api_offers:
        key = normalize_offer_id(o.offer_id)
        if key in merged:
            # Match — mark as "both", enrich with API data
            existing = merged[key]
            existing.source = "both"
            if o.price and not existing.price:
                existing.price = o.price
            if o.mileage and not existing.mileage:
                existing.mileage = o.mileage
            if o.year and not existing.year:
                existing.year = o.year
        else:
            merged[key] = o

    return list(merged.values())


# ── Format ────────────────────────────────────────────────────────

SOURCE_ICON = {
    "email": "📧",
    "api": "🔌",
    "both": "📧+🔌",
}


def format_offers(offers):
    """Format merged offers into Telegram message text."""
    not_purchased = [o for o in offers if o.category == "not_purchased"]
    back_on_sale = [o for o in offers if o.category == "back_on_sale"]

    parts = []

    if not_purchased:
        lines = [f"🔍 Не выкупленные: {len(not_purchased)} авто\n"]
        for o in not_purchased:
            icon = SOURCE_ICON.get(o.source, "")
            extra = _format_extra(o)
            lines.append(f"{o.offer_id}/")
            lines.append(f"{o.salon} used {o.brand} {o.model}{extra} {icon}")
            lines.append(o.mobile_url)
        parts.append("\n".join(lines))

    if back_on_sale:
        lines = [f"🔄 Снова в продаже: {len(back_on_sale)} авто\n"]
        for o in back_on_sale:
            icon = SOURCE_ICON.get(o.source, "")
            extra = _format_extra(o)
            lines.append(f"{o.offer_id}/")
            lines.append(f"{o.salon} used {o.brand} {o.model}{extra} {icon}")
            lines.append(o.mobile_url)
        parts.append("\n".join(lines))

    return "\n\n".join(parts)


def _format_extra(o):
    """Format price/mileage suffix if available from API."""
    parts = []
    if o.price:
        if o.price >= 1_000_000:
            parts.append(f"{o.price / 1_000_000:.1f}М₽".replace(".0М", "М"))
        else:
            parts.append(f"{o.price:,}₽".replace(",", " "))
    if o.mileage:
        parts.append(f"{o.mileage:,}km".replace(",", " "))
    if parts:
        return " | " + " | ".join(parts)
    return ""


# ── Telegram ─────────────────────────────────────────────────────

def send_telegram(text):
    if not TELEGRAM_CHAT_ID:
        print(text)
        return False

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    chunks = []
    while len(text) > 4000:
        split_pos = text.rfind("\n", 0, 4000)
        if split_pos == -1:
            split_pos = 4000
        chunks.append(text[:split_pos])
        text = text[split_pos:].lstrip()
    chunks.append(text)

    for chunk in chunks:
        data = json.dumps({
            "chat_id": TELEGRAM_CHAT_ID,
            "text": chunk,
        }).encode("utf-8")

        req = urllib.request.Request(
            url, data=data,
            headers={"Content-Type": "application/json"},
        )
        try:
            with urllib.request.urlopen(req, timeout=10) as resp:
                result = json.loads(resp.read().decode())
                if not result.get("ok"):
                    print(f"[TG] Error: {result}")
                    return False
        except Exception as e:
            print(f"[TG] Error: {e}")
            return False

    return True


# ── Main ─────────────────────────────────────────────────────────

def run():
    print(f"[{datetime.now().strftime('%H:%M')}] Email Digest Bot (Email + API)")

    state = load_state()

    # 1. Fetch emails → parse Excel → email_offers
    emails, new_email_ids = fetch_today_emails(state)
    email_offers = []
    downloaded_files = []

    for subject, files in emails:
        for filepath in files:
            downloaded_files.append(filepath)
            fname = os.path.basename(filepath).lower()

            if "не_выкупленные" in fname or "не выкупленные" in fname:
                email_offers.extend(parse_not_purchased(filepath))
            elif "снова_в_продаже" in fname or "снова в продаже" in fname:
                email_offers.extend(parse_back_on_sale(filepath))
            else:
                print(f"[Email] Unknown format: {fname}")

    print(f"[Email] Parsed {len(email_offers)} offers from email")

    # 2. Fetch API comeback → api_offers
    api_offers = fetch_api_comeback()

    # 3. Filter out already-sent API offer_ids
    sent_api_ids = set(state.get("api_offer_ids", []))
    if sent_api_ids:
        before = len(api_offers)
        api_offers = [o for o in api_offers if normalize_offer_id(o.offer_id) not in sent_api_ids]
        skipped = before - len(api_offers)
        if skipped:
            print(f"[API] Filtered {skipped} already-sent offers")

    # 4. Merge
    merged = merge_offers(email_offers, api_offers)

    if not merged and not emails:
        print("No new data from email or API.")
        save_state(state)
        return

    if not merged:
        print("Emails found but no parseable offers.")
        # Still update email IDs so we don't re-process
        state["email_message_ids"] = list(set(state["email_message_ids"]) | new_email_ids)
        save_state(state)
        return

    # 5. Format
    has_api = any(o.source in ("api", "both") for o in merged)
    source_label = "Email + API" if has_api else "Email"
    header = f"📧 Дайджест auto.ru — {datetime.now().strftime('%d.%m.%Y')} ({source_label})\n"

    body = format_offers(merged)
    message = header + "\n" + body

    # 6. Send
    success = send_telegram(message)

    # 7. Save state
    if success:
        state["email_message_ids"] = list(set(state["email_message_ids"]) | new_email_ids)
        new_api_ids = {normalize_offer_id(o.offer_id) for o in api_offers}
        state["api_offer_ids"] = list(set(state.get("api_offer_ids", [])) | new_api_ids)
        state["api_last_fetch"] = datetime.now().isoformat()
        save_state(state)

    print(f"{'✅ Sent' if success else '⚠️ Failed'} — {len(merged)} offers ({source_label})")

    # 8. Cleanup Excel files
    for f in downloaded_files:
        try:
            os.remove(f)
        except OSError:
            pass


if __name__ == "__main__":
    run()
