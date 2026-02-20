"""Fetch latest Excel attachments from Gmail via IMAP."""
import imaplib
import email
import os
from email.header import decode_header
from dotenv import load_dotenv

load_dotenv()

GMAIL_USER = os.getenv("GMAIL_USER")
GMAIL_APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD")
IMAP_SERVER = "imap.gmail.com"
SENDER_FILTER = "no-reply@business.auto.ru"
DOWNLOAD_DIR = os.path.join(os.path.dirname(__file__), "downloads")

os.makedirs(DOWNLOAD_DIR, exist_ok=True)


def decode_subject(msg):
    subject, encoding = decode_header(msg["Subject"])[0]
    if isinstance(subject, bytes):
        subject = subject.decode(encoding or "utf-8")
    return subject


def fetch_latest_attachments(max_emails=5):
    """Connect to Gmail, find emails from sender, download Excel attachments."""
    print(f"Connecting to {IMAP_SERVER} as {GMAIL_USER}...")
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(GMAIL_USER, GMAIL_APP_PASSWORD)
    mail.select("inbox")

    # Search for emails from the sender
    status, messages = mail.search(None, f'(FROM "{SENDER_FILTER}")')
    if status != "OK" or not messages[0]:
        print("No emails found from", SENDER_FILTER)
        mail.logout()
        return []

    email_ids = messages[0].split()
    print(f"Found {len(email_ids)} emails from {SENDER_FILTER}")

    # Take the latest ones
    latest_ids = email_ids[-max_emails:]
    downloaded = []

    for eid in reversed(latest_ids):
        status, data = mail.fetch(eid, "(RFC822)")
        if status != "OK":
            continue

        msg = email.message_from_bytes(data[0][1])
        subject = decode_subject(msg)
        date = msg["Date"]
        print(f"\n--- Email: {subject} ({date}) ---")

        for part in msg.walk():
            content_type = part.get_content_type()
            filename = part.get_filename()

            if filename:
                # Decode filename if needed
                decoded_name, enc = decode_header(filename)[0]
                if isinstance(decoded_name, bytes):
                    decoded_name = decoded_name.decode(enc or "utf-8")

                print(f"  Attachment: {decoded_name} ({content_type})")

                # Save Excel files
                if decoded_name.endswith((".xlsx", ".xls", ".csv")):
                    filepath = os.path.join(DOWNLOAD_DIR, decoded_name)
                    with open(filepath, "wb") as f:
                        f.write(part.get_payload(decode=True))
                    print(f"  -> Saved to {filepath}")
                    downloaded.append(filepath)

    mail.logout()
    print(f"\nTotal downloaded: {len(downloaded)} files")
    return downloaded


if __name__ == "__main__":
    files = fetch_latest_attachments()
    if files:
        # Quick peek at structure
        try:
            import openpyxl
            for f in files:
                if f.endswith(".xlsx"):
                    wb = openpyxl.load_workbook(f, read_only=True)
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        print(f"\n=== {os.path.basename(f)} / {sheet_name} ===")
                        for i, row in enumerate(ws.iter_rows(values_only=True)):
                            if i < 10:  # First 10 rows
                                print(row)
                            else:
                                break
                        print(f"Total rows: {ws.max_row}, columns: {ws.max_column}")
                    wb.close()
        except ImportError:
            print("Install openpyxl: pip install openpyxl")
